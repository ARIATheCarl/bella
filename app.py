import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
import math
import calendar
from FinMind.data import DataLoader
# 將極值標記改為紅色／藍色邊框樣式
from openpyxl.styles import Border, Side

st.set_page_config(page_title="蘇文政股期分析專用工具", layout="centered")

# ===== 下載 FinMind 股票資訊 =====
@st.cache_data
def get_stock_info():
    api = DataLoader()
    return api.taiwan_stock_info()

stock_info_df = get_stock_info()


st.title("蘇文政股期分析專用工具")

interval = st.radio("選擇統計區間", ["日", "週", "月"], horizontal=True)


# 1. 先產生 tuple 清單
stock_options = [
    (row['stock_id'], row['stock_name'], row['type'], row['date'])
    for _, row in stock_info_df.iterrows()
]
# 股票選單
if not stock_options:
    st.error("查無可用股票資料，請確認 FinMind API 是否正常。")
    st.stop()
# 2. 做一份漂亮的顯示清單
display_options = [
    f"{row[0]:>6} {row[1]:<8}"
    for row in stock_options
]
# 3. selectbox 用 index 選
selected_index = st.selectbox(
    "選擇股票代碼", 
    range(len(stock_options)), 
    format_func=lambda i: display_options[i]
)
stock_id, stock_name, stock_type, stock_date = stock_options[selected_index]

# 取得更完整的資訊
row = stock_info_df[stock_info_df["stock_id"] == stock_id].iloc[0]
industry = row.get("industry_category", "未知產業")
market = row.get("type", "未知市場")
listed_date = row.get("date", "未知日期")

# 顯示更多細節
st.info(
    f"""
    **{stock_name}**（代碼：{stock_id}）
    - 市場別：{market}
    - 產業類別：{industry}
    - 資料更新日：{listed_date}
    """
)

# 如果想要當日收盤價等資訊，可以加上即時查詢（如有即時 API）


# 年月日下拉式選單
years = list(range(1991, 2036))
months = list(range(1, 13))
days = list(range(1, 32))

col1, col2, col3 = st.columns(3)
with col1:
    start_year = st.selectbox("起始年", years, index=years.index(datetime.today().year))
with col2:
    start_month = st.selectbox("起始月", months, index=0)
with col3:
    start_day = st.selectbox("起始日", days, index=0)

col4, col5, col6 = st.columns(3)
with col4:
    end_year = st.selectbox("結束年", years, index=years.index(datetime.today().year), key="end_year")
with col5:
    end_month = st.selectbox("結束月", months, index=datetime.today().month-1, key="end_month")
with col6:
    end_day = st.selectbox("結束日", days, index=datetime.today().day-1, key="end_day")

# 處理不合法日期（例如2/30）
try:
    start_date = datetime(start_year, start_month, start_day)
except ValueError:
    start_date = datetime(start_year, start_month, 1)
    st.warning("起始日設為該月1日（選擇的日期無效）")
try:
    end_date = datetime(end_year, end_month, end_day)
except ValueError:
    end_date = datetime(end_year, end_month, 1)
    st.warning("結束日設為該月1日（選擇的日期無效）")

# ====== FinMind 取資料 ======
def fetch_finmind_data(stock_id: str, start: str, end: str) -> list:
    api = DataLoader()
    api.login_by_token(api_token="eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJkYXRlIjoiMjAyNS0wNi0wMyAxMDozMzoxNSIsInVzZXJfaWQiOiJjYXJsNzk5MSIsImlwIjoiNDkuMjE0LjAuMTQxIn0.Qzdlv5fe2J3rRUCpAYDltguY_oGgLlqp7kwILmnTVdA")  # <<<<<< 請填入你的 token
    df = api.taiwan_stock_daily(
        stock_id=stock_id,
        start_date=start,
        end_date=end,
    )
    if df.empty:
        st.error(f"FinMind 查不到 {stock_id} 的資料，請檢查代碼或日期。")
        st.stop()
    df["date"] = pd.to_datetime(df["date"])
    if "max_price" in df.columns:
        high_col = "max_price"
        low_col = "min_price"
    elif "max" in df.columns:
        high_col = "max"
        low_col = "min"
    else:
        st.error(f"FinMind回傳資料找不到max_price/max欄位，欄位有：{df.columns}")
        st.stop()
    class StockData:
        def __init__(self, date, high, low, capacity):
            self.date = date
            self.high = high
            self.low = low
            self.capacity = capacity
    return [
        StockData(row["date"], row[high_col], row[low_col], row["Trading_Volume"])
        for _, row in df.iterrows()
    ]

# ========== 主程式 ==========
if st.button("產生報表"):
    start_date = datetime.combine(start_date, datetime.min.time())
    end_date = datetime.combine(end_date, datetime.min.time())
    today = datetime.today()

    # 區間對齊
    if interval == "週":
        start_date = start_date - timedelta(days=start_date.weekday())
        end_date = end_date + timedelta(days=6 - end_date.weekday())
    elif interval == "月":
        start_date = start_date.replace(day=1)
        last_day = calendar.monthrange(end_date.year, end_date.month)[1]
        end_date = end_date.replace(day=last_day)
    if end_date > today:
        end_date = today

    # 主資料
    raw_data = fetch_finmind_data(stock_id, start_date.strftime("%Y-%m-%d"), end_date.strftime("%Y-%m-%d"))
    filtered = [d for d in raw_data if start_date <= d.date <= end_date]
    if not filtered:
        st.error("查無資料，請檢查股票代碼與時間範圍。")
        st.stop()

    # 比對基準
    def fetch_prev(interval, stock_id, start_date, end_date):
        if interval == "週":
            prev_start = start_date - timedelta(days=7)
            prev_end = start_date - timedelta(days=1)
            raw_prev = fetch_finmind_data(stock_id, prev_start.strftime("%Y-%m-%d"), prev_end.strftime("%Y-%m-%d"))
            prev_filtered = [d for d in raw_prev if prev_start <= d.date <= prev_end]
            if prev_filtered:
                prev_high = max(d.high for d in prev_filtered)
                prev_low = min(d.low for d in prev_filtered)
                prev_volume = sum(d.capacity for d in prev_filtered)
                prev_diff = prev_high - prev_low
            else:
                prev_high, prev_low, prev_volume, prev_diff = None, None, None, None
        elif interval == "月":
            prev_month_end = start_date - timedelta(days=1)
            prev_month_start = prev_month_end.replace(day=1)
            raw_prev = fetch_finmind_data(stock_id, prev_month_start.strftime("%Y-%m-%d"), prev_month_end.strftime("%Y-%m-%d"))
            prev_filtered = [d for d in raw_prev if prev_month_start <= d.date <= prev_month_end]
            if prev_filtered:
                prev_high = max(d.high for d in prev_filtered)
                prev_low = min(d.low for d in prev_filtered)
                prev_volume = sum(d.capacity for d in prev_filtered)
                prev_diff = prev_high - prev_low
            else:
                prev_high, prev_low, prev_volume, prev_diff = None, None, None, None
        else:
            extra_date = start_date - timedelta(days=14)
            raw_prev = fetch_finmind_data(stock_id, extra_date.strftime("%Y-%m-%d"), start_date.strftime("%Y-%m-%d"))
            prev_filtered = [d for d in raw_prev if d.date < start_date]
            if prev_filtered:
                d = max(prev_filtered, key=lambda x: x.date)
                prev_high = d.high
                prev_low = d.low
                prev_volume = d.capacity
                prev_diff = prev_high - prev_low
            else:
                prev_high, prev_low, prev_volume, prev_diff = None, None, None, None
        return prev_high, prev_low, prev_volume, prev_diff

    prev_high, prev_low, prev_volume, prev_diff = fetch_prev(interval, stock_id, start_date, end_date)

    # 轉 DataFrame
    df = pd.DataFrame([{
        '日期': d.date,
        '最高價': d.high,
        '最低價': d.low,
        '成交量': d.capacity
    } for d in filtered])

    # 統計區間
    if interval == "日":
        agg_df = df.copy()
    elif interval == "週":
        df["iso_year"] = df["日期"].apply(lambda x: x.isocalendar()[0])
        df["iso_week"] = df["日期"].apply(lambda x: x.isocalendar()[1])
        grouped = df.groupby(["iso_year", "iso_week"])
        agg_df = grouped.agg({
            "日期": "last",
            "最高價": "max",
            "最低價": "min",
            "成交量": "sum"
        }).reset_index(drop=True)
    else:  # 月
        df["month"] = df["日期"].apply(lambda x: x.month)
        df["year"] = df["日期"].apply(lambda x: x.year)
        grouped = df.groupby(["year", "month"])
        agg_df = grouped.agg({
            "日期": "last",
            "最高價": "max",
            "最低價": "min",
            "成交量": "sum"
        }).reset_index(drop=True)
        
    # 加上是否為極值欄位（修正週與月的分組）
    agg_df["是否最高點"] = False
    agg_df["是否最低點"] = False
    
    if interval == "日":
        df["週"] = df["日期"].apply(lambda x: x.isocalendar()[1])
        df["年"] = df["日期"].apply(lambda x: x.isocalendar()[0])
        groups = df.groupby(["年", "週"])
    elif interval == "週":
        agg_df["年"] = agg_df["日期"].dt.year
        agg_df["月"] = agg_df["日期"].dt.month
        groups = agg_df.groupby(["年", "月"])
    else:  # 月
        agg_df["年"] = agg_df["日期"].dt.year
        groups = agg_df.groupby("年")
    
    for name, group in groups:
        max_high = group["最高價"].max()
        min_low = group["最低價"].min()
        # 最高價第一個
        max_date = group[group["最高價"] == max_high].sort_values("日期").iloc[-1]["日期"]
        # 最低價最後一個
        min_date = group[group["最低價"] == min_low].sort_values("日期").iloc[-1]["日期"]
        agg_df.loc[agg_df["日期"] == max_date, "是否最高點"] = True
        agg_df.loc[agg_df["日期"] == min_date, "是否最低點"] = True


    # 顏色計算
    agg_df["差色"] = ""
    agg_df["高色"] = ""
    agg_df["低色"] = ""
    agg_df["成交符"] = ""
    agg_df["符色"] = ""

    for i, row in agg_df.iterrows():
        if i == 0 and prev_high is not None:
            cmp_high = prev_high
            cmp_low = prev_low
            cmp_volume = prev_volume
            cmp_diff = prev_diff
        else:
            prev_row = agg_df.iloc[i-1]
            cmp_high = prev_row["最高價"]
            cmp_low = prev_row["最低價"]
            cmp_volume = prev_row["成交量"]
            cmp_diff = prev_row["最高價"] - prev_row["最低價"]

        diff = row["最高價"] - row["最低價"]
        agg_df.loc[i, "高色"] = "FFCC3333" if row["最高價"] >= cmp_high else "FF3366CC"
        agg_df.loc[i, "低色"] = "FFCC3333" if row["最低價"] >= cmp_low else "FF3366CC"
        agg_df.loc[i, "成交符"] = "■"
        agg_df.loc[i, "符色"] = "FFCC3333" if row["成交量"] >= cmp_volume else "FF3366CC"
        agg_df.loc[i, "差色"] = "FFCC3333" if diff >= cmp_diff else "FF3366CC"

    # 分頁
    chunk_size = 43
    chunks = []
    for i in range(0, len(agg_df), chunk_size):
        chunks.append(agg_df.iloc[i:i+chunk_size].reset_index(drop=True))

    blocks_per_sheet = 3
    total_blocks = len(chunks)
    total_pages = math.ceil(total_blocks / blocks_per_sheet)

    wb = Workbook()
    ws = wb.active
    ws.title = f"{interval}報表"

    title = f"{stock_name} {start_date.strftime('%Y-%m-%d')}～{end_date.strftime('%Y-%m-%d')}（{interval}）"
    sheet_count = 0
    headers = ["日期", "", "高", "低", "漲幅", "量", ""] * 3
    starts = [1, 8, 15]
    prev_month = None
    prev_year = None

    for block, data in enumerate(chunks):
        if block % blocks_per_sheet == 0:
            sheet_count = block // blocks_per_sheet
            ws = wb.active if block == 0 else wb.create_sheet(title=f"{interval}報表{sheet_count+1}")

            ws.freeze_panes = "A4"
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 1
            ws.page_setup.scale = None
            ws.page_setup.orientation = "portrait"
            ws.page_setup.paperSize = 11
            ws.sheet_properties = WorksheetProperties(pageSetUpPr=PageSetupProperties(fitToPage=True))
            ws.page_setup.horizontalCentered = True
            ws.page_setup.verticalCentered = True
            ws.page_margins = PageMargins(
                left=0.1, right=0.1,
                top=0.1, bottom=0.1,
                header=0.0, footer=0.0
            )
    
            ws.insert_rows(1)
            ws.insert_rows(2)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=20)
            page_info = f"（第 {sheet_count+1}/{total_pages} 頁）" if total_pages > 1 else ""
            title_cell = ws.cell(row=1, column=1, value=f"{title}{page_info}")
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 30
            for i, h in enumerate(headers):
                cell = ws.cell(row=2, column=i + 1, value=h)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            for i in starts:
                ws.merge_cells(start_row=2, start_column=i, end_row=2, end_column=i+1)
            sheet_count += 1

        col = starts[block % 3]
        row_index = 3

        for i, row in data.iterrows():
            dt = row["日期"]
            if interval == "日":
                if (prev_month != None) and (dt.month != prev_month):
                    day_str = f"{dt.month}/{dt.day}"
                else:
                    day_str = f"{dt.day}"
                week_str = ["一", "二", "三", "四", "五", "六", "日"][dt.weekday()]
                prev_month = dt.month
                ws.cell(row=row_index, column=col, value=day_str).alignment = Alignment(horizontal="center")
                ws.cell(row=row_index, column=col+1, value=week_str).alignment = Alignment(horizontal="center")
            elif interval == "週":
                if (prev_year != None) and (dt.year != prev_year):
                    day_str = f"{dt.year}/{dt.month}/{dt.day}"
                else:
                    day_str = f"{dt.month}/{dt.day}"
                prev_year = dt.year
                ws.merge_cells(start_row=row_index, start_column=col, end_row=row_index, end_column=col+1)
                ws.cell(row=row_index, column=col, value=day_str).alignment = Alignment(horizontal="center")
            else:
                if (prev_year != None) and (dt.year != prev_year):
                    day_str = f"{dt.year}/{dt.month}"
                else:
                    day_str = f"{dt.month}"
                prev_year = dt.year
                ws.merge_cells(start_row=row_index, start_column=col, end_row=row_index, end_column=col+1)
                ws.cell(row=row_index, column=col, value=day_str).alignment = Alignment(horizontal="center")

            # 定義邊框樣式
            red_border = Border(outline=True, left=Side(style="thin", color="009A72"),
                                right=Side(style="thin", color="009A72"),
                                top=Side(style="thin", color="009A72"),
                                bottom=Side(style="thin", color="009A72"))
            blue_border = Border(outline=True, left=Side(style="thin", color="009A72"),
                                 right=Side(style="thin", color="009A72"),
                                 top=Side(style="thin", color="009A72"),
                                 bottom=Side(style="thin", color="009A72"))

            # 修改寫入 Excel 的最高價與最低價欄位樣式：
            h = ws.cell(row=row_index, column=col+2, value=row["最高價"])
            if row.get("是否最高點", False):
                h.font = Font(color=row["高色"], bold=True)
                h.border = red_border
            else:
                h.font = Font(color=row["高色"])
            h.alignment = Alignment(horizontal="center")
            
            l = ws.cell(row=row_index, column=col+3, value=row["最低價"])
            if row.get("是否最低點", False):
                l.font = Font(color=row["低色"], bold=True)
                l.border = blue_border
            else:
                l.font = Font(color=row["低色"])
            l.alignment = Alignment(horizontal="center")

            d_value = round(row["最高價"] - row["最低價"], 2)
            d = ws.cell(row=row_index, column=col+4, value=d_value)
            d.font = Font(color=row["差色"])
            d.alignment = Alignment(horizontal="center")

            v = ws.cell(row=row_index, column=col+5, value=row["成交符"])
            v.font = Font(color=row["符色"], size=10)
            v.alignment = Alignment(horizontal="center")
            row_index += 1

        # 欄寬自動
        for col_cells in ws.iter_cols(min_row=3, max_col=ws.max_column, max_row=ws.max_row):
            col_letter = get_column_letter(col_cells[0].column)
            max_len = max(len(str(c.value)) if c.value else 0 for c in col_cells)
            ws.column_dimensions[col_letter].width = max(2, min(max_len + 2, 14))
        for i in range(8, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(i)].width = ws.column_dimensions[get_column_letter(i-7)].width


    buffer = BytesIO()
    wb.save(buffer)
    st.success("✅ 報表產出成功")
    st.download_button("下載 Excel", data=buffer.getvalue(), file_name=f"{title}.xlsx")
